from __future__ import annotations

from pathlib import Path
from typing import List, Optional

import openpyxl

from app.models import ProductInput

COLUMN_MAP = {
    "상품명": "product_name",
    "레퍼런스": "product_code",
    "색상": "color",
    "소재": "material",
    "카테고리": "category",
    "가격": "price",
    "사이즈": "size",
    "설명": "description",
}


def brand_name_from_path(excel_path: Path) -> str:
    return excel_path.stem


def load_products_from_excel(excel_path: Path) -> List[ProductInput]:
    brand_name = brand_name_from_path(excel_path)
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    products = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        product = _row_to_product(dict(zip(headers, row_values)), brand_name)
        if product is not None:
            products.append(product)

    wb.close()
    return products


def _row_to_product(row: dict, brand_name: str) -> Optional[ProductInput]:
    kwargs: dict = {"brand_name": brand_name}
    for excel_col, field in COLUMN_MAP.items():
        val = row.get(excel_col)
        if val is not None and str(val).strip():
            kwargs[field] = val

    if not kwargs.get("product_code") and not kwargs.get("product_name"):
        return None

    try:
        return ProductInput(**kwargs)
    except Exception:
        return None
