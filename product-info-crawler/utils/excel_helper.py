import re
from pathlib import Path
from typing import Optional


EXCEL_OUTPUT_FOLDER_NAME = "Brand_Product_Data"
SHEET_HEADERS = ["번호", "상품명", "레퍼런스", "색상", "카테고리", "가격"]
WRITE_MODE_APPEND = "append"
WRITE_MODE_OVERWRITE = "overwrite"
INVALID_FILE_NAME_CHARS = r'[<>:"/\\|?*]'


def sanitize_excel_file_name(name: str) -> str:
    sanitized_name = re.sub(INVALID_FILE_NAME_CHARS, "_", name).strip().strip(".")
    return sanitized_name or "brand"


def sanitize_excel_sheet_name(name: str) -> str:
    sanitized_name = re.sub(r"[:\\/?*\[\]]", "_", name).strip()
    return sanitized_name[:31] or "Sheet1"


def get_excel_output_path(brand_name: str, desktop_dir: Optional[Path] = None) -> Path:
    desktop_path = desktop_dir or (Path.home() / "Desktop")
    output_dir = desktop_path / EXCEL_OUTPUT_FOLDER_NAME
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_brand_name = sanitize_excel_file_name(brand_name)
    return output_dir / f"{safe_brand_name}.xlsx"


def get_or_create_excel_workbook(file_path: Path):
    from openpyxl import Workbook, load_workbook

    if file_path.exists():
        return load_workbook(file_path)

    return Workbook()


def _is_empty_default_sheet(workbook) -> bool:
    if len(workbook.sheetnames) != 1:
        return False

    worksheet = workbook.active
    return worksheet.title == "Sheet" and worksheet.max_row == 1 and worksheet["A1"].value is None


def get_or_create_excel_worksheet(workbook, sheet_name: str):
    sanitized_name = sanitize_excel_sheet_name(sheet_name)

    if sanitized_name in workbook.sheetnames:
        return workbook[sanitized_name], False

    if _is_empty_default_sheet(workbook):
        worksheet = workbook.active
        worksheet.title = sanitized_name
        return worksheet, True

    return workbook.create_sheet(title=sanitized_name), True


def ensure_excel_headers(worksheet) -> None:
    header_row = [cell.value for cell in worksheet[1]]
    if header_row[: len(SHEET_HEADERS)] != SHEET_HEADERS:
        worksheet.delete_rows(1, worksheet.max_row)
        worksheet.append(SHEET_HEADERS)


def clear_excel_worksheet(worksheet) -> None:
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(SHEET_HEADERS)


def append_products_to_excel_sheet(
    worksheet,
    products: list[dict],
    write_mode: str = WRITE_MODE_APPEND,
) -> int:
    if write_mode == WRITE_MODE_OVERWRITE:
        clear_excel_worksheet(worksheet)
    else:
        ensure_excel_headers(worksheet)

    if not products:
        return 0

    start_row = worksheet.max_row if write_mode == WRITE_MODE_APPEND else 1
    for idx, product in enumerate(products, start=start_row):
        worksheet.append(
            [idx, product["name"], product["reference"], product["colors"], product["category"], product["price"]]
        )

    return len(products)


def save_excel_workbook(workbook, file_path: Path) -> None:
    workbook.save(file_path)
