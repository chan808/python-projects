import random
import time
from pathlib import Path
from typing import Optional

import streamlit as st
import undetected_chromedriver as uc
from webdriver_manager.chrome import ChromeDriverManager

from scrapers.registry import get_scraper_class
from utils.config_loader import ConfigError, discover_brand_configs, load_brand_config
from utils.excel_helper import (
    EXCEL_OUTPUT_FOLDER_NAME,
    append_products_to_excel_sheet,
    get_excel_output_path,
    get_or_create_excel_workbook,
    get_or_create_excel_worksheet,
    save_excel_workbook,
)
from utils.gsheet_helper import (
    WRITE_MODE_APPEND,
    WRITE_MODE_OVERWRITE,
    append_products_to_sheet,
    get_or_create_worksheet,
    get_spreadsheet,
)


PROJECT_ROOT = Path(__file__).resolve().parent
CONFIG_DIR = PROJECT_ROOT / "config"
STORAGE_TARGET_GOOGLE_SHEETS = "google_sheets"
STORAGE_TARGET_EXCEL = "excel"

# 재고 확인 API가 구현된 브랜드 목록 — 새 브랜드 추가 시 여기에 등록
STOCK_CHECK_SUPPORTED_BRANDS = {"lv"}


# ── 상품 수집 ────────────────────────────────────────────────────────────────

def build_driver(config: dict) -> uc.Chrome:
    chrome_options = uc.ChromeOptions()
    scraping_settings = config.get("scraping_settings", {})

    if scraping_settings.get("headless", False):
        chrome_options.add_argument("--headless")

    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--lang=ko-KR")
    chrome_options.add_argument("--start-maximized")

    # webdriver_manager로 Chrome 버전에 맞는 드라이버를 자동 설치
    # undetected_chromedriver가 드라이버를 패치해 Akamai 봇 탐지를 우회
    driver_path = ChromeDriverManager().install()
    driver = uc.Chrome(
        options=chrome_options,
        driver_executable_path=driver_path,
        use_subprocess=True,
    )

    driver.set_window_size(1920, 1080)
    time.sleep(random.uniform(1, 2))

    return driver


def save_products(
    storage_target: str,
    brand_name: str,
    category_name: str,
    write_mode: str,
    spreadsheet,
    workbook,
    excel_path: Optional[Path],
    products: list[dict],
) -> tuple[int, str]:
    if storage_target == STORAGE_TARGET_GOOGLE_SHEETS:
        sheet_name = f"{brand_name} : {category_name}"
        worksheet, created = get_or_create_worksheet(spreadsheet, sheet_name)
        saved_count = append_products_to_sheet(worksheet, products, write_mode=write_mode)

        if created:
            detail_message = "새 구글 시트 생성"
        elif write_mode == WRITE_MODE_OVERWRITE:
            detail_message = "기존 구글 시트 덮어쓰기"
        else:
            detail_message = "기존 구글 시트 이어쓰기"

        return saved_count, detail_message

    worksheet, created = get_or_create_excel_worksheet(workbook, category_name)
    saved_count = append_products_to_excel_sheet(worksheet, products, write_mode=write_mode)
    save_excel_workbook(workbook, excel_path)

    if created:
        detail_message = "새 엑셀 시트 생성"
    elif write_mode == WRITE_MODE_OVERWRITE:
        detail_message = "기존 엑셀 시트 덮어쓰기"
    else:
        detail_message = "기존 엑셀 시트 이어쓰기"

    return saved_count, detail_message


def run_scraping(
    brand_name: str,
    scraper_key: str,
    config: dict,
    write_mode: str,
    selected_categories: list[str],
    storage_target: str,
) -> None:
    driver = None
    spreadsheet = None
    workbook = None
    excel_path = None

    try:
        config["project_root"] = str(PROJECT_ROOT)
        scraper_cls = get_scraper_class(scraper_key)
        if getattr(scraper_cls, "requires_driver", True):
            driver = build_driver(config)

        scraper = scraper_cls(driver, config)

        if storage_target == STORAGE_TARGET_GOOGLE_SHEETS:
            google_sheets_config = config["google_sheets"]
            spreadsheet = get_spreadsheet(
                google_sheets_config["service_account_file"],
                google_sheets_config["spreadsheet_name"],
            )
        else:
            excel_path = get_excel_output_path(brand_name)
            workbook = get_or_create_excel_workbook(excel_path)

        categories_to_process = [
            (category_name, url)
            for category_name, url in config["categories"].items()
            if category_name in selected_categories
        ]

        progress_bar = st.progress(0.0, text="크롤링 준비 중")
        status_container = st.container()
        result_container = st.container()
        category_results = []
        inter_category_delay_sec = float(config.get("scraping_settings", {}).get("inter_category_delay_sec", 0))

        for index, (category_name, url) in enumerate(categories_to_process, start=1):
            progress_ratio = index / len(categories_to_process)
            progress_bar.progress(
                progress_ratio,
                text=f"{index}/{len(categories_to_process)} {category_name} 처리 중",
            )

            try:
                with status_container:
                    st.info(f"{category_name} 수집 중")

                products = scraper.parse_category(category_name, url)

                saved_count, detail_message = save_products(
                    storage_target=storage_target,
                    brand_name=brand_name,
                    category_name=category_name,
                    write_mode=write_mode,
                    spreadsheet=spreadsheet,
                    workbook=workbook,
                    excel_path=excel_path,
                    products=products,
                )

                scraping_settings = config.get("scraping_settings", {})
                if scraping_settings.get("fetch_detail", False) and products:
                    detail_status = status_container.empty()
                    checkpoint_warnings: list[str] = []

                    def _on_detail_progress(current: int, total: int, name: str) -> None:
                        detail_status.caption(f"상세정보 수집 중: {current}/{total} — {name[:40]}")

                    def _on_checkpoint(current_products: list[dict]) -> None:
                        try:
                            save_products(
                                storage_target=storage_target,
                                brand_name=brand_name,
                                category_name=category_name,
                                write_mode=WRITE_MODE_OVERWRITE,
                                spreadsheet=spreadsheet,
                                workbook=workbook,
                                excel_path=excel_path,
                                products=current_products,
                            )
                        except PermissionError as exc:
                            checkpoint_warnings.append(str(exc))

                    products, enrich_fail = scraper.enrich_with_details(
                        products,
                        on_progress=_on_detail_progress,
                        on_checkpoint=_on_checkpoint,
                    )
                    detail_status.empty()

                    saved_count, detail_message = save_products(
                        storage_target=storage_target,
                        brand_name=brand_name,
                        category_name=category_name,
                        write_mode=WRITE_MODE_OVERWRITE,
                        spreadsheet=spreadsheet,
                        workbook=workbook,
                        excel_path=excel_path,
                        products=products,
                    )
                    if enrich_fail > 0:
                        detail_message += f" (상세정보 실패 {enrich_fail}건 — tmp_debug 폴더 확인)"
                    for warning in checkpoint_warnings:
                        with result_container:
                            st.warning(warning)

                category_results.append(
                    {
                        "category": category_name,
                        "status": "success",
                        "saved_count": saved_count,
                        "detail": detail_message,
                    }
                )
            except Exception as exc:
                category_results.append(
                    {
                        "category": category_name,
                        "status": "error",
                        "saved_count": 0,
                        "detail": str(exc),
                    }
                )

            if inter_category_delay_sec > 0 and index < len(categories_to_process):
                actual_delay = inter_category_delay_sec * random.uniform(0.8, 1.5)
                with status_container:
                    st.caption(f"다음 카테고리를 위해 {actual_delay:.1f}초간 대기 중...")
                time.sleep(actual_delay)

        progress_bar.progress(1.0, text="크롤링 완료")

        success_count = sum(1 for result in category_results if result["status"] == "success")
        error_count = len(category_results) - success_count

        with result_container:
            if error_count == 0:
                st.success(f"{brand_name} 전체 카테고리 수집이 끝났습니다. 성공 {success_count}건")
            else:
                st.warning(
                    f"{brand_name} 수집이 완료되었습니다. 성공 {success_count}건, 실패 {error_count}건"
                )

            for result in category_results:
                if result["status"] == "success":
                    st.success(
                        f"{result['category']} 완료: {result['saved_count']}건 저장 ({result['detail']})"
                    )
                else:
                    st.error(f"{result['category']} 실패: {result['detail']}")

            if storage_target == STORAGE_TARGET_EXCEL and excel_path is not None:
                st.info(f"엑셀 저장 위치: {excel_path} ({EXCEL_OUTPUT_FOLDER_NAME} 폴더)")

    finally:
        if driver is not None:
            driver.quit()


# ── 재고 매장 확인 ───────────────────────────────────────────────────────────

def run_inventory_check(
    excel_path: Path,
    selected_sheets: list[str],
    delay_sec: float,
    config: dict,
) -> None:
    from openpyxl import load_workbook
    from utils.lv_stock_api import fetch_stores_with_stock_via_driver
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By

    try:
        workbook = load_workbook(excel_path)
    except PermissionError:
        st.error(f"엑셀 파일이 열려 있습니다. 닫은 후 다시 시도하세요: {excel_path.name}")
        return

    tasks = []
    for sheet_name in selected_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        for row in workbook[sheet_name].iter_rows(min_row=2):
            sku = str(row[2].value or "").strip()
            if sku:
                tasks.append((row, sku))

    if not tasks:
        st.warning("선택한 시트에 레퍼런스(SKU)가 있는 상품이 없습니다.")
        return

    total = len(tasks)
    progress_bar = st.progress(0.0, text="브라우저 세션 초기화 중...")
    status = st.empty()

    driver = None
    try:
        driver = build_driver(config)

        # LV 도메인 접속 — Akamai 쿠키(_abck) 획득
        driver.get("https://kr.louisvuitton.com/kor-kr/homepage")
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(random.uniform(4, 6))

        success_count = 0
        fail_count = 0

        for i, (row, sku) in enumerate(tasks):
            category = str(row[5].value or "")
            status.caption(f"{category} — {sku} ({i + 1}/{total})")

            try:
                stores = fetch_stores_with_stock_via_driver(driver, sku)
                row[9].value = ", ".join(stores)
                success_count += 1
            except Exception:
                fail_count += 1

            progress_bar.progress((i + 1) / total, text=f"{i + 1}/{total} 처리 중")

            if (i + 1) % 20 == 0:
                try:
                    save_excel_workbook(workbook, excel_path)
                except PermissionError:
                    pass

            if i + 1 < total:
                time.sleep(delay_sec)

    finally:
        if driver:
            driver.quit()

    try:
        save_excel_workbook(workbook, excel_path)
    except PermissionError as exc:
        st.error(str(exc))
        return

    progress_bar.progress(1.0, text="완료")

    if fail_count == 0:
        st.success(f"완료: {success_count}개 처리. 저장 위치: {excel_path}")
    else:
        st.warning(f"완료: 성공 {success_count}개, 실패 {fail_count}개. 저장 위치: {excel_path}")


# ── UI ───────────────────────────────────────────────────────────────────────

def _scraping_mode_ui(brand_summaries: list) -> None:
    brand_by_id = {brand.brand_id: brand for brand in brand_summaries}
    selected_brand_id = st.radio(
        "브랜드 선택",
        [brand.brand_id for brand in brand_summaries],
        format_func=lambda brand_id: brand_by_id[brand_id].display_name,
        horizontal=True,
    )
    selected_brand = brand_by_id[selected_brand_id]

    storage_target = st.radio(
        "저장 대상",
        [STORAGE_TARGET_GOOGLE_SHEETS, STORAGE_TARGET_EXCEL],
        format_func=lambda target: "Google Sheets" if target == STORAGE_TARGET_GOOGLE_SHEETS else "Excel",
        horizontal=True,
    )
    write_mode = st.radio(
        "저장 방식",
        [WRITE_MODE_OVERWRITE, WRITE_MODE_APPEND],
        format_func=lambda mode: "덮어쓰기" if mode == WRITE_MODE_OVERWRITE else "이어쓰기",
        horizontal=True,
    )

    try:
        config = load_brand_config(
            selected_brand.config_path,
            project_root=PROJECT_ROOT,
            validate_credentials=(storage_target == STORAGE_TARGET_GOOGLE_SHEETS),
        )
    except ConfigError as exc:
        st.error(f"설정 파일을 불러오지 못했습니다: {exc}")
        st.stop()

    category_names = list(config["categories"].keys())
    selected_categories = st.multiselect(
        "수집할 카테고리",
        category_names,
        default=category_names,
        help="선택한 카테고리만 실행합니다.",
    )

    if storage_target == STORAGE_TARGET_EXCEL:
        excel_path = get_excel_output_path(selected_brand.display_name)
        st.caption(f"엑셀 파일: {excel_path}")

    if not selected_categories:
        st.warning("최소 1개 이상의 카테고리를 선택해야 합니다.")

    if st.button("크롤링 시작", type="primary"):
        if not selected_categories:
            st.stop()
        st.info(f"{selected_brand.display_name} 수집을 시작합니다.")
        try:
            run_scraping(
                brand_name=selected_brand.display_name,
                scraper_key=selected_brand.scraper_key,
                config=config,
                write_mode=write_mode,
                selected_categories=selected_categories,
                storage_target=storage_target,
            )
        except Exception as exc:
            st.exception(exc)


def _inventory_mode_ui(brand_summaries: list) -> None:
    from openpyxl import load_workbook

    eligible_brands = [b for b in brand_summaries if b.scraper_key in STOCK_CHECK_SUPPORTED_BRANDS]

    if not eligible_brands:
        st.info("현재 재고 확인을 지원하는 브랜드가 없습니다.")
        return

    brand_by_id = {b.brand_id: b for b in eligible_brands}
    selected_id = st.radio(
        "브랜드 선택",
        [b.brand_id for b in eligible_brands],
        format_func=lambda bid: brand_by_id[bid].display_name,
        horizontal=True,
    )
    selected_brand = brand_by_id[selected_id]
    excel_path = get_excel_output_path(selected_brand.display_name)

    if not excel_path.exists():
        st.warning("엑셀 파일이 없습니다. 먼저 '상품 정보 수집'을 실행하세요.")
        st.caption(f"예상 경로: {excel_path}")
        return

    wb = load_workbook(excel_path, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    st.caption(f"파일: {excel_path}")

    selected_sheets = st.multiselect(
        "확인할 카테고리 (시트)",
        sheet_names,
        default=list(sheet_names),
    )
    delay_sec = st.slider(
        "요청 간격 (초)",
        min_value=0.5, max_value=5.0, value=1.0, step=0.5,
        help="간격이 짧을수록 빠르지만 API 차단 위험이 있습니다.",
    )

    if not selected_sheets:
        st.warning("최소 1개 이상의 카테고리를 선택해야 합니다.")

    try:
        config = load_brand_config(selected_brand.config_path, project_root=PROJECT_ROOT, validate_credentials=False)
    except ConfigError as exc:
        st.error(f"설정 파일을 불러오지 못했습니다: {exc}")
        return

    if st.button("재고 확인 시작", type="primary", disabled=not selected_sheets):
        run_inventory_check(excel_path, selected_sheets, delay_sec, config)


def main() -> None:
    st.title("멀티 브랜드 상품 크롤러")

    try:
        brand_summaries = discover_brand_configs(CONFIG_DIR, PROJECT_ROOT)
    except ConfigError as exc:
        st.error(f"브랜드 설정을 불러오지 못했습니다: {exc}")
        st.stop()

    mode = st.radio(
        "실행 모드",
        ["상품 정보 수집", "재고 매장 확인"],
        horizontal=True,
    )
    st.divider()

    if mode == "재고 매장 확인":
        _inventory_mode_ui(brand_summaries)
    else:
        _scraping_mode_ui(brand_summaries)


if __name__ == "__main__":
    main()
