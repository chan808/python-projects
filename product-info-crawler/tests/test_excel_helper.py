import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from utils.excel_helper import (
    EXCEL_OUTPUT_FOLDER_NAME,
    WRITE_MODE_APPEND,
    WRITE_MODE_OVERWRITE,
    append_products_to_excel_sheet,
    get_excel_output_path,
    get_or_create_excel_worksheet,
    sanitize_excel_file_name,
    sanitize_excel_sheet_name,
)

SAMPLE_PRODUCTS = [
    {"name": "Wallet A", "reference": "REF-001", "colors": "Black", "category": "Wallet(Man)", "price": 1000000},
    {"name": "Wallet B", "reference": "REF-002", "colors": "Brown", "category": "Wallet(Man)", "price": 1200000},
]


class ExcelHelperTest(unittest.TestCase):
    def test_get_excel_output_path_creates_desktop_output_folder(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            desktop_dir = Path(temp_dir) / "Desktop"
            desktop_dir.mkdir()

            output_path = get_excel_output_path("Celine", desktop_dir=desktop_dir)

            self.assertEqual(output_path.name, "Celine.xlsx")
            self.assertEqual(output_path.parent.name, EXCEL_OUTPUT_FOLDER_NAME)
            self.assertTrue(output_path.parent.exists())

    def test_sanitize_excel_file_name_replaces_invalid_characters(self) -> None:
        self.assertEqual(sanitize_excel_file_name('Brand:/\\"?*<>|'), "Brand_________")

    def test_sanitize_excel_sheet_name_limits_length_and_invalid_characters(self) -> None:
        sheet_name = sanitize_excel_sheet_name("Shoes:/[Woman] 2026 Spring Collection Very Long Name")

        self.assertNotIn(":", sheet_name)
        self.assertLessEqual(len(sheet_name), 31)

    def test_overwrite_mode_replaces_all_rows_and_numbers_from_1(self) -> None:
        wb = Workbook()
        ws, _ = get_or_create_excel_worksheet(wb, "Test")

        append_products_to_excel_sheet(ws, SAMPLE_PRODUCTS, write_mode=WRITE_MODE_OVERWRITE)
        append_products_to_excel_sheet(ws, SAMPLE_PRODUCTS, write_mode=WRITE_MODE_OVERWRITE)

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 2)
        self.assertEqual(data_rows[0][0], 1)
        self.assertEqual(data_rows[1][0], 2)

    def test_append_mode_continues_row_numbering(self) -> None:
        wb = Workbook()
        ws, _ = get_or_create_excel_worksheet(wb, "Test")

        append_products_to_excel_sheet(ws, SAMPLE_PRODUCTS, write_mode=WRITE_MODE_APPEND)
        append_products_to_excel_sheet(ws, SAMPLE_PRODUCTS, write_mode=WRITE_MODE_APPEND)

        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(data_rows), 4)
        self.assertEqual(data_rows[0][0], 1)
        self.assertEqual(data_rows[2][0], 3)

    def test_empty_products_returns_zero(self) -> None:
        wb = Workbook()
        ws, _ = get_or_create_excel_worksheet(wb, "Test")

        count = append_products_to_excel_sheet(ws, [], write_mode=WRITE_MODE_APPEND)
        self.assertEqual(count, 0)

    def test_product_fields_written_in_correct_columns(self) -> None:
        wb = Workbook()
        ws, _ = get_or_create_excel_worksheet(wb, "Test")

        append_products_to_excel_sheet(ws, [SAMPLE_PRODUCTS[0]], write_mode=WRITE_MODE_OVERWRITE)

        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        # 번호, 상품명, 레퍼런스, 색상, 소재, 카테고리, 가격
        self.assertEqual(row[1], "Wallet A")
        self.assertEqual(row[2], "REF-001")
        self.assertEqual(row[3], "Black")
        self.assertEqual(row[4], "")        # 소재 (SAMPLE_PRODUCTS에 없음)
        self.assertEqual(row[5], "Wallet(Man)")
        self.assertEqual(row[6], 1000000)


if __name__ == "__main__":
    unittest.main()
